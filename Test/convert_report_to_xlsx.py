import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font

def parse_report(file_path):
    with open(file_path, 'r') as f:
        lines = f.readlines()

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"

    # Define the Cambria font styles
    cambria_font = Font(name='Cambria')
    cambria_bold_font = Font(name='Cambria', bold=True)

    # Set header row with Cambria bold font
    headers = ["File", "Line", "Test Name", "Result", "Details"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cambria_bold_font

    # Regex to match test results
    test_result_pattern = re.compile(r'(?P<file>.*?):(?P<line>\d+):(?P<test_name>[^:]+):(?P<result>PASS|FAIL)(?::(?P<details>.*))?')

    for line in lines:
        match = test_result_pattern.match(line)
        if match:
            row = [match.group("file"), match.group("line"), match.group("test_name"), match.group("result"), match.group("details") or ""]
            ws.append(row)
            # Apply Cambria font to each cell in the row
            for cell in ws[ws.max_row]:
                cell.font = cambria_font

    # Add summary section at the end
    summary_result = lines[-1].strip() if lines else "No summary available"
    summary_text = f"Summary: {summary_result}"

    # Append a row for summary and merge cells
    ws.append([""] * 5)  # Add an empty row for the summary
    summary_row_index = ws.max_row
    ws.merge_cells(start_row=summary_row_index, start_column=1, end_row=summary_row_index, end_column=5)
    summary_cell = ws.cell(row=summary_row_index, column=1, value=summary_text)
    summary_cell.font = cambria_font

    return wb

def main():
    if len(sys.argv) != 2:
        print("Usage: python convert_report_to_xlsx.py <input_file>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = input_file.replace('.txt', '.xlsx')

    wb = parse_report(input_file)
    wb.save(output_file)
    print(f"XLSX REPORT GENERATED: {output_file}")

if __name__ == "__main__":
    main()
